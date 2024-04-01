"""Excel report generation using openpyxl."""
import io

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


HEADER_FILL = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def _style_header(ws, row=1, max_col=10):
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")
        cell.border = BORDER


def generate_candidates_excel(candidates: list[dict]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Candidates"

    headers = ["ID", "Name", "Email", "Phone", "Job", "Experience (yrs)", "Status", "Created"]
    ws.append(headers)
    _style_header(ws, max_col=len(headers))

    for c in candidates:
        ws.append([
            c.get("id", ""),
            c.get("full_name", ""),
            c.get("email", ""),
            c.get("phone", ""),
            c.get("job_title", ""),
            c.get("experience_years", ""),
            c.get("status", ""),
            c.get("created_at", ""),
        ])

    for col in ws.columns:
        max_length = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_length + 4, 40)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


def generate_evaluation_excel(evaluations: list[dict]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Evaluations"

    headers = [
        "ID", "Candidate", "Job", "Communication", "Technical",
        "Confidence", "Domain Knowledge", "Problem Solving",
        "Overall", "AI Recommendation", "HR Decision", "Date",
    ]
    ws.append(headers)
    _style_header(ws, max_col=len(headers))

    for e in evaluations:
        ws.append([
            e.get("id", ""),
            e.get("candidate_name", ""),
            e.get("job_title", ""),
            e.get("communication_score", 0),
            e.get("technical_score", 0),
            e.get("confidence_score", 0),
            e.get("domain_knowledge_score", 0),
            e.get("problem_solving_score", 0),
            e.get("overall_score", 0),
            e.get("ai_recommendation", ""),
            e.get("hr_decision", ""),
            e.get("evaluated_at", ""),
        ])

    for col in ws.columns:
        max_length = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_length + 4, 40)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()
